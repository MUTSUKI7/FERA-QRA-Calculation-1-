# -*- coding: utf-8 -*-
"""
QRA Contour Viewer — Streamlit Cloud 배포용 앱

기존 로컬 스크립트(thermal_contour_code_260814.py, LSIR_contour_code_260814.py)를
웹앱 형태로 리팩터링한 버전입니다.

핵심 변경점 (로컬 스크립트 대비):
  1. 하드코딩된 파일 경로(C:\\Users\\...) 대신 st.file_uploader로 엑셀/이미지를 받음
  2. 전역 상수(BACK_FRAC, CROSS_FRAC 등)를 사이드바 슬라이더로 노출 -> 실시간 조정 가능
  3. 파일 I/O(디스크 저장) 대신 메모리(BytesIO)에서 처리 후 st.image/st.pyplot으로 표시
  4. 매번 재계산하지 않도록 @st.cache_data로 엑셀 파싱 결과를 캐싱

실행: streamlit run app.py
배포: GitHub 리포에 이 파일 + requirements.txt + (선택) packages.txt 올린 뒤
      share.streamlit.io 에서 리포 연결
"""

import io
import math

import cv2
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.colors import LogNorm
from matplotlib.lines import Line2D
from PIL import Image, ImageDraw, ImageFont

import streamlit as st

st.set_page_config(page_title="QRA Contour Viewer", layout="wide")
st.title("QRA Contour 자동 생성기")

# ------------------------------------------------------------------
# 사이드바: 입력 파일 + 파라미터
# ------------------------------------------------------------------
with st.sidebar:
    st.header("1. 입력 파일")
    excel_file = st.file_uploader(
        "QRA_Master_Template.xlsx", type=["xlsx"],
        help="IS_Coordinates / PHAST_Distances / Wind rose(Meteorology) 등의 시트를 포함한 엑셀"
    )
    image_file = st.file_uploader(
        "배경 위성사진 (선택)", type=["png", "jpg", "jpeg"],
        help="업로드하지 않으면 격자선만 있는 빈 배경을 사용합니다"
    )

    st.header("2. 격자 설정")
    grid_extent_m = st.number_input("전체 범위 (m)", value=1000.0, step=50.0)
    cell_size_m = st.number_input("격자 한 칸 크기 (m)", value=50.0, step=10.0)
    canvas_px = st.slider("캔버스 해상도 (px, 이미지 미업로드시)", 400, 1200, 800, 50)

    st.header("3. Contour 종류")
    contour_type = st.radio("표시할 결과", ["Thermal (풍향별 lobe, 윤곽선)", "LSIR (개인위험도, heatmap+contour)"])

    st.header("4. Lobe 형태 가정치")
    st.caption("엑셀에 없는 임의 가정값 — 실측/PHAST 방향별 출력과 다를 수 있습니다")
    back_frac = st.slider("풍상측 비율 (BACK_FRAC)", 0.0, 1.0, 0.30, 0.05)
    cross_frac = st.slider("풍직각 반폭 비율 (CROSS_FRAC)", 0.0, 1.0, 0.35, 0.05)

    if contour_type.startswith("LSIR"):
        st.header("5. LSIR 계산 설정")
        grid_resolution_m = st.slider("계산 해상도 (m/격자점, 작을수록 정밀하지만 느림)", 1, 20, 5, 1)
        run_lsir = st.button("LSIR 계산 실행", type="primary")

if excel_file is None:
    st.info("왼쪽에서 QRA_Master_Template.xlsx 파일을 먼저 업로드하세요.")
    st.stop()

EXCEL_BYTES = excel_file.getvalue()


# ------------------------------------------------------------------
# 공통 유틸
# ------------------------------------------------------------------
def make_canvas(size_px, grid_extent_m, cell_size_m):
    """배경 이미지가 없을 때 격자선이 그려진 빈 캔버스 생성"""
    img = Image.new("RGB", (size_px, size_px), (235, 235, 235))
    draw = ImageDraw.Draw(img)
    n_cells = int(round(grid_extent_m / cell_size_m))
    for i in range(n_cells + 1):
        p = int(round(i * size_px / n_cells))
        draw.line([(p, 0), (p, size_px)], fill=(200, 200, 200), width=1)
        draw.line([(0, p), (size_px, p)], fill=(200, 200, 200), width=1)
    return img


def load_font(size=11):
    for path in [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
    ]:
        try:
            return ImageFont.truetype(path, size)
        except Exception:
            continue
    return ImageFont.load_default()


@st.cache_data(show_spinner=False)
def load_thermal_data(excel_bytes):
    xls = pd.ExcelFile(io.BytesIO(excel_bytes))
    is_coords = pd.read_excel(xls, sheet_name="IS_Coordinates")
    is_coords["Loc_X_1000"] = pd.to_numeric(is_coords["Loc_X_1000"], errors="coerce")
    is_coords["Loc_Y_1000"] = pd.to_numeric(is_coords["Loc_Y_1000"], errors="coerce")

    phast = pd.read_excel(xls, sheet_name="PHAST_Distances")
    for c in ["Jet_12.5", "Jet_37.5", "Pool_12.5", "Pool_37.5"]:
        if c in phast.columns:
            phast[c] = pd.to_numeric(phast[c], errors="coerce")

    wind_rose = pd.read_excel(xls, sheet_name="Wind rose(Meteorology)")

    by_weather = (
        phast.groupby(["IS_ID", "Weather_Class"])[["Jet_12.5", "Jet_37.5", "Pool_12.5", "Pool_37.5"]]
        .max()
        .reset_index()
    )

    def fire_type_for(is_id):
        sub = by_weather[by_weather["IS_ID"] == is_id]
        has_jet = sub[["Jet_12.5", "Jet_37.5"]].notna().any().any()
        has_pool = sub[["Pool_12.5", "Pool_37.5"]].notna().any().any()
        return "Jet fire" if has_jet else ("Pool fire" if has_pool else None)

    fire_types = {is_id: fire_type_for(is_id) for is_id in is_coords["IS_ID"]}
    points_m = {row["IS_ID"]: (row["Loc_X_1000"], row["Loc_Y_1000"]) for _, row in is_coords.iterrows()}

    return by_weather, wind_rose, fire_types, points_m


# ------------------------------------------------------------------
# Thermal contour (풍향별 lobe)
# ------------------------------------------------------------------
def lobe_polygon_px(center_m, R_m, bearing_from_deg, back_frac, cross_frac, px_per_m, H, n=48):
    downwind_deg = (bearing_from_deg + 180.0) % 360.0
    rad = math.radians(downwind_deg)
    dx_down, dy_down = math.sin(rad), math.cos(rad)
    dx_cross, dy_cross = math.cos(rad), -math.sin(rad)

    semi_major = R_m * (1 + back_frac) / 2.0
    semi_minor = R_m * cross_frac
    offset = R_m * (1 - back_frac) / 2.0

    ellipse_cx = center_m[0] + dx_down * offset
    ellipse_cy = center_m[1] + dy_down * offset

    theta = np.linspace(0, 2 * np.pi, n, endpoint=False)
    local_x = semi_major * np.cos(theta)
    local_y = semi_minor * np.sin(theta)
    east = ellipse_cx + local_x * dx_down + local_y * dx_cross
    north = ellipse_cy + local_x * dy_down + local_y * dy_cross

    def m_to_px(x_m, y_m):
        return x_m * px_per_m, H - y_m * px_per_m

    return [m_to_px(e, n_) for e, n_ in zip(east, north)]


def build_union_mask(level_col, size, by_weather, wind_rose, fire_types, points_m,
                      back_frac, cross_frac, px_per_m):
    mask = np.zeros((size[1], size[0]), dtype=np.uint8)
    for _, wr_row in wind_rose.iterrows():
        wc = wr_row["Weather_Class"]
        bearing = wr_row["Angle_degree"]
        matches = by_weather[by_weather["Weather_Class"] == wc]
        for _, row in matches.iterrows():
            is_id = row["IS_ID"]
            ftype = fire_types.get(is_id)
            if ftype is None or is_id not in points_m:
                continue
            col = level_col.replace("LEVEL", "Jet" if ftype == "Jet fire" else "Pool")
            R = row[col]
            if pd.isna(R) or R <= 0:
                continue
            poly = lobe_polygon_px(points_m[is_id], R, bearing, back_frac, cross_frac,
                                    px_per_m, size[1])
            poly_int = np.array([[int(round(x)), int(round(y))] for x, y in poly], dtype=np.int32)
            cv2.fillPoly(mask, [poly_int], 255, lineType=cv2.LINE_AA)
    return mask


def mask_to_contours(mask):
    contours, _ = cv2.findContours(mask, cv2.RETR_CCOMP, cv2.CHAIN_APPROX_SIMPLE)
    return contours


def draw_outline_contours(draw, contours, color, width):
    for cnt in contours:
        if len(cnt) < 2:
            continue
        pts = [(int(p[0][0]), int(p[0][1])) for p in cnt]
        pts.append(pts[0])
        draw.line(pts, fill=color, width=width, joint="curve")


def render_thermal(excel_bytes, image_file, grid_extent_m, cell_size_m, canvas_px,
                    back_frac, cross_frac):
    by_weather, wind_rose, fire_types, points_m = load_thermal_data(excel_bytes)

    if image_file is not None:
        img = Image.open(image_file).convert("RGB")
    else:
        img = make_canvas(canvas_px, grid_extent_m, cell_size_m)
    W, H = img.size
    px_per_m = W / grid_extent_m

    def m_to_px(x_m, y_m):
        return x_m * px_per_m, H - y_m * px_per_m

    mask_125 = build_union_mask("LEVEL_12.5", img.size, by_weather, wind_rose, fire_types,
                                 points_m, back_frac, cross_frac, px_per_m)
    mask_375 = build_union_mask("LEVEL_37.5", img.size, by_weather, wind_rose, fire_types,
                                 points_m, back_frac, cross_frac, px_per_m)
    contours_125 = mask_to_contours(mask_125)
    contours_375 = mask_to_contours(mask_375)

    overlay = Image.new("RGBA", img.size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(overlay)
    COLOR_125 = (255, 200, 0, 255)
    COLOR_375 = (220, 20, 20, 255)
    draw_outline_contours(draw, contours_125, COLOR_125, width=3)
    draw_outline_contours(draw, contours_375, COLOR_375, width=3)

    font_small = load_font(11)
    for is_id, (x_m, y_m) in points_m.items():
        ftype = fire_types.get(is_id)
        if ftype is None:
            continue
        cx, cy = m_to_px(x_m, y_m)
        draw.ellipse([cx - 4, cy - 4, cx + 4, cy + 4], fill=(255, 255, 255, 255),
                     outline=(0, 0, 0, 255), width=1)
        draw.text((cx + 6, cy - 14), f"{is_id} [{ftype}]", fill=(255, 255, 255, 255),
                   font=font_small, stroke_width=2, stroke_fill=(0, 0, 0, 255))

    combined = Image.alpha_composite(img.convert("RGBA"), overlay).convert("RGB")
    return combined


# ------------------------------------------------------------------
# LSIR (개인위험도) — LSIR_contour_code_260814.py 로직 포팅
# ------------------------------------------------------------------
LIQUID_IS_IDS = {"IS02_FC02", "IS33_FC41"}
HOLE_SIZES = [2, 5.4, 22.3, 86.6, 150]


def to_float_or_none(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().upper()
    if s in ("NR", "NA", ""):
        return None
    try:
        return float(v)
    except ValueError:
        return None


@st.cache_data(show_spinner=False)
def load_lsir_data(excel_bytes):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)

    is_coords = {}
    for row in wb["IS_Coordinates"].iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        is_coords[row[0]] = (float(row[1]), float(row[2]))

    leak_freq = {}
    ws = wb["Leak_Frequencies"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        is_id = row[0]
        leak_freq[is_id] = {}
        for hs, val in zip(HOLE_SIZES, row[1:6]):
            leak_freq[is_id][hs] = float(val) if val is not None else 0.0

    phast = {}
    for row in wb["PHAST_Distances"].iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        is_id, hole_size, weather_class = row[0], row[1], row[2]
        phast[(is_id, float(hole_size), weather_class)] = {
            "Jet_12.5": to_float_or_none(row[3]), "Jet_37.5": to_float_or_none(row[4]),
            "Pool_12.5": to_float_or_none(row[5]), "Pool_37.5": to_float_or_none(row[6]),
            "Flash_LFL": to_float_or_none(row[7]),
            "Exp_0.3": to_float_or_none(row[8]), "Exp_0.5": to_float_or_none(row[9]),
        }

    wind_rows_raw, total_prob = [], 0.0
    for row in wb["Wind rose(Meteorology)"].iter_rows(min_row=2, values_only=True):
        if row[0] is None or row[3] is None:
            continue
        wind_rows_raw.append({"angle_deg": float(row[1]), "weather_class": row[2], "prob_raw": float(row[3])})
        total_prob += float(row[3])
    wind_rows = [{**r, "prob": r["prob_raw"] / total_prob} for r in wind_rows_raw]

    vuln = {}
    for row in wb["Vulnerability_Criteria"].iter_rows(min_row=2, values_only=True):
        vuln[(row[0], row[1])] = float(row[3])
    vuln_map = {
        "thermal_37.5": vuln.get(("Thermal (Jet/Pool)", 37.5)),
        "thermal_12.5": vuln.get(("Thermal (Jet/Pool)", 12.5)),
        "flash": vuln.get(("Flash Fire", "5 (Methane LFL)")),
        "exp_0.5": vuln.get(("Overpressure", 0.5)),
        "exp_0.3": vuln.get(("Overpressure", 0.3)),
    }

    ignition = {}
    for row in wb["Ignition_Probability"].iter_rows(min_row=3, values_only=True):
        if row[0] is None or not isinstance(row[1], (int, float)):
            continue
        ignition[(row[0], float(row[1]))] = {
            "P_jet": row[6] or 0.0, "P_flash": row[7] or 0.0,
            "P_vce": row[8] or 0.0, "P_pool": row[9] or 0.0,
        }

    return is_coords, leak_freq, phast, wind_rows, vuln_map, ignition


def circle_mask(X, Y, center, r):
    if r is None:
        return None
    return np.sqrt((X - center[0]) ** 2 + (Y - center[1]) ** 2) <= r


def lobe_mask(X, Y, center, r, bearing_from_deg, back_frac, cross_frac):
    if r is None or r <= 0:
        return None
    downwind_deg = (bearing_from_deg + 180.0) % 360.0
    rad = math.radians(downwind_deg)
    dx_down, dy_down = math.sin(rad), math.cos(rad)
    dx_cross, dy_cross = math.cos(rad), -math.sin(rad)
    semi_major = r * (1 + back_frac) / 2.0
    semi_minor = r * cross_frac
    offset = r * (1 - back_frac) / 2.0
    ecx = center[0] + dx_down * offset
    ecy = center[1] + dy_down * offset
    rel_x, rel_y = X - ecx, Y - ecy
    along = rel_x * dx_down + rel_y * dy_down
    cross = rel_x * dx_cross + rel_y * dy_cross
    return (along / semi_major) ** 2 + (cross / semi_minor) ** 2 <= 1.0


def add_band(risk_field, mask_severe, mask_mild, v_severe, v_mild, weight):
    if weight <= 0:
        return
    base = 0.0
    if mask_mild is not None:
        risk_field[mask_mild] += weight * v_mild
        base = v_mild
    if mask_severe is not None:
        risk_field[mask_severe] += weight * (v_severe - base)


def add_single(risk_field, mask, v, weight):
    if weight <= 0 or mask is None or v is None:
        return
    risk_field[mask] += weight * v


def build_lsir_field(is_coords, leak_freq, phast, wind_rows, vuln, ignition,
                      X, Y, back_frac, cross_frac):
    risk_field = np.zeros_like(X, dtype=float)
    for is_id, center in is_coords.items():
        is_liquid = is_id in LIQUID_IS_IDS
        for hs in HOLE_SIZES:
            freq = leak_freq.get(is_id, {}).get(hs, 0.0)
            if freq <= 0:
                continue
            ign = ignition.get((is_id, hs))
            if ign is None:
                continue
            for wr in wind_rows:
                wc, bearing, wind_p = wr["weather_class"], wr["angle_deg"], wr["prob"]
                d = phast.get((is_id, hs, wc))
                if d is None:
                    continue
                if is_liquid:
                    w = freq * wind_p * ign["P_pool"]
                    ms = lobe_mask(X, Y, center, d["Pool_37.5"], bearing, back_frac, cross_frac)
                    mm = lobe_mask(X, Y, center, d["Pool_12.5"], bearing, back_frac, cross_frac)
                    add_band(risk_field, ms, mm, vuln["thermal_37.5"], vuln["thermal_12.5"], w)
                else:
                    w_jet = freq * wind_p * ign["P_jet"]
                    ms = lobe_mask(X, Y, center, d["Jet_37.5"], bearing, back_frac, cross_frac)
                    mm = lobe_mask(X, Y, center, d["Jet_12.5"], bearing, back_frac, cross_frac)
                    add_band(risk_field, ms, mm, vuln["thermal_37.5"], vuln["thermal_12.5"], w_jet)

                    w_flash = freq * wind_p * ign["P_flash"]
                    mf = lobe_mask(X, Y, center, d["Flash_LFL"], bearing, back_frac, cross_frac)
                    add_single(risk_field, mf, vuln["flash"], w_flash)

                    w_vce = freq * wind_p * ign["P_vce"]
                    mse = circle_mask(X, Y, center, d["Exp_0.5"])
                    mmi = circle_mask(X, Y, center, d["Exp_0.3"])
                    add_band(risk_field, mse, mmi, vuln["exp_0.5"], vuln["exp_0.3"], w_vce)
    return risk_field


def render_lsir(excel_bytes, image_file, grid_extent_m, resolution_m, back_frac, cross_frac):
    is_coords, leak_freq, phast, wind_rows, vuln, ignition = load_lsir_data(excel_bytes)

    n = int(grid_extent_m / resolution_m) + 1
    xs = np.linspace(0, grid_extent_m, n)
    ys = np.linspace(0, grid_extent_m, n)
    X, Y = np.meshgrid(xs, ys)

    risk_field = build_lsir_field(is_coords, leak_freq, phast, wind_rows, vuln, ignition,
                                   X, Y, back_frac, cross_frac)

    fig, ax = plt.subplots(figsize=(8, 8), dpi=130)
    if image_file is not None:
        img = plt.imread(image_file)
        ax.imshow(img, extent=[0, grid_extent_m, 0, grid_extent_m], origin="upper")

    plot_field = np.clip(risk_field, 1e-12, None)
    levels = [1e-6, 1e-5, 1e-4, 1e-3]
    colors = ["#2ca02c", "#ffff33", "#ff9900", "#ff0000"]
    ax.contour(X, Y, plot_field, levels=levels, norm=LogNorm(), colors=colors, linewidths=1.8)

    handles = [Line2D([0], [0], color=c, lw=2) for c in colors]
    labels = [f"{lv:.0e} /yr" for lv in levels]
    ax.legend(handles, labels, loc="lower left", fontsize=8, framealpha=0.85, title="LSIR level")

    for is_id, (x, y) in is_coords.items():
        ax.plot(x, y, "o", color="deeppink", markersize=5)
        ax.annotate(is_id, (x, y), fontsize=6, color="black",
                    textcoords="offset points", xytext=(4, 4))

    ax.set_xlim(0, grid_extent_m)
    ax.set_ylim(0, grid_extent_m)
    ax.set_xlabel("X (m)")
    ax.set_ylabel("Y (m)")
    ax.set_title("LSIR Contour (1/year)")
    fig.tight_layout()
    return fig, risk_field.max()


# ------------------------------------------------------------------
# 메인 렌더링
# ------------------------------------------------------------------
if contour_type.startswith("Thermal"):
    try:
        result_img = render_thermal(EXCEL_BYTES, image_file, grid_extent_m, cell_size_m,
                                     canvas_px, back_frac, cross_frac)
        st.image(result_img, use_container_width=True)

        buf = io.BytesIO()
        result_img.save(buf, format="PNG")
        st.download_button("PNG 다운로드", buf.getvalue(), "thermal_contour.png", "image/png")
    except KeyError as e:
        st.error(f"엑셀에서 필요한 시트/열을 찾을 수 없습니다: {e}")
else:
    if not run_lsir:
        st.info("왼쪽 사이드바에서 파라미터를 설정한 뒤 '[LSIR 계산 실행]' 버튼을 누르세요. "
                 "(격자점마다 반복 계산이라 시간이 걸릴 수 있습니다)")
    else:
        try:
            with st.spinner("LSIR 필드 계산 중..."):
                fig, max_risk = render_lsir(EXCEL_BYTES, image_file, grid_extent_m,
                                             grid_resolution_m, back_frac, cross_frac)
            st.pyplot(fig)
            st.caption(f"Max LSIR ≈ {max_risk:.3e} /yr")

            buf = io.BytesIO()
            fig.savefig(buf, format="png", dpi=150)
            st.download_button("PNG 다운로드", buf.getvalue(), "lsir_contour.png", "image/png")
        except KeyError as e:
            st.error(f"엑셀에서 필요한 시트/열을 찾을 수 없습니다: {e}")
